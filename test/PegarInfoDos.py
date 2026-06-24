import gc
import sys
import time
import uuid
import locale
import threading
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from filelock import FileLock
import streamlit as st
from datetime import datetime

RUTA_CSS = r"O:\Gerencia Contraloria\Analitica Contraloria\Automatiaciones Ambiente Pruebas\Carpeta Miguel Cardona\FormularioAreas\styles\STYLES.css"
RUTA_ICON = r"O:\Gerencia Contraloria\Analitica Contraloria\Automatiaciones Ambiente Pruebas\Carpeta Miguel Cardona\FormularioAreas\images\gco_ico.svg"
RUTA_IMAGE = r"O:\Gerencia Contraloria\Analitica Contraloria\Automatiaciones Ambiente Pruebas\Carpeta Miguel Cardona\FORMULARIOS\images\GContraloria.png"
RUTA_ARCHIVO = r"O:\Gerencia Contraloria\Analitica Contraloria\Automatiaciones Ambiente Pruebas\Carpeta Miguel Cardona\FormularioAreas\input\Ingreso Datos Informe Gerencia Contraloria - Eficiencias y Volumetria.xlsm"
RUTA_ICON_MARCAS = r"O:\Gerencia Contraloria\Analitica Contraloria\Automatiaciones Ambiente Pruebas\Carpeta Miguel Cardona\FormularioAreas\images\footer_marcas.svg"
RUTA_BIENVENIDA_GCO = r"O:\Gerencia Contraloria\Analitica Contraloria\Automatiaciones Ambiente Pruebas\Carpeta Miguel Cardona\FormularioAreas\images\bienvenidaGCO.png"
RUTA_LOCK = RUTA_ARCHIVO + ".lock"

locale.setlocale(locale.LC_ALL, '')
sys.stdout.reconfigure(encoding="utf-8")

CREDENCIALES = {
    "Miguel Cardona":   ["mcardona",  "777"],
    "Jorge Herrera":    ["jorgeeh",   "1212"],
    "Alberto Cortés":   ["albertoc",  "2323"],
    "Oscar Yepes":      ["oscardy",   "3434"],
    "Dora Gómez":       ["doragc",    "4545"],
    "Zaneida Restrepo": ["zrestrepo", "5656"],
    "Ana Romero":       ["anamr",     "6767"],
}

AREAS = {
    "mcardona":  "Admin",
    "jorgeeh":   "Analítica de Contraloría",
    "albertoc":  "Control de Operaciones",
    "oscardy":   "Administrativa",
    "doragc":    "Riesgos y Cumplimiento",
    "zrestrepo": "Impuestos",
    "anamr":     "Contabilidad",
}

MESES = [
    "Enero", "Febrero", "Marzo", "Abril",
    "Mayo", "Junio", "Julio", "Agosto",
    "Septiembre", "Octubre", "Noviembre", "Diciembre",
]

CONTRASENA_EXCEL = "54312"

# ============================================ CARGA PARÁMETROS ====================================

@st.cache_data
def cargar_excel():
    return pd.read_excel(
        io=RUTA_ARCHIVO,
        sheet_name="Parámetros",
        skiprows=1,
    )

df = cargar_excel()

# ============================================ FUNCTIONS ==========================================

def sincronizar_filas(conceptos_seleccionados):
    """Sincroniza la lista de filas con los conceptos seleccionados en el multiselect."""

    # Eliminar filas de conceptos que fueron deseleccionados
    st.session_state.filas = [
        f for f in st.session_state.filas
        if f["concepto"] in conceptos_seleccionados
    ]

    # Si venimos de eliminar una fila manualmente, no recrear filas
    if st.session_state.saltarse_sync:
        st.session_state.saltarse_sync = False
        return

    # Crear fila base solo para conceptos nuevos (que no existan ya)
    existentes = {f["concepto"] for f in st.session_state.filas}

    for concepto in conceptos_seleccionados:
        if concepto not in existentes:
            st.session_state.filas.append({
                "concepto": concepto,
                "id": str(uuid.uuid4())
            })


def parametros(area):
    """Retorna las listas de parámetros según el área del usuario autenticado."""

    lista_especificaciones = df.iloc[:, 14].dropna().drop_duplicates().tolist()
    lista_ciudades         = df.iloc[:, 16].dropna().drop_duplicates().tolist()

    if area in ("Analítica de Contraloría", "Admin"):
        col_concepto_nuevo = 22
        col_unidad         = 20
    elif area == "Control de Operaciones":
        col_concepto_nuevo = 38
        col_unidad         = 36
    elif area == "Administrativa":
        col_concepto_nuevo = 10
        col_unidad         = 8
    elif area == "Riesgos y Cumplimiento":
        col_concepto_nuevo = 50
        col_unidad         = 52
    elif area == "Impuestos":
        col_concepto_nuevo = 46
        col_unidad         = 44
    elif area == "Contabilidad":
        col_concepto_nuevo = 30
        col_unidad         = 28

    df_filtrado = df[df.iloc[:, col_concepto_nuevo].notna()].copy()

    lista_concepto_nuevo  = df_filtrado.iloc[:, col_concepto_nuevo].drop_duplicates().tolist()
    diccionario_unidades  = dict(zip(
        df_filtrado.iloc[:, col_concepto_nuevo],
        df_filtrado.iloc[:, col_unidad]
    ))

    return lista_especificaciones, lista_ciudades, lista_concepto_nuevo, diccionario_unidades


# ========================== ESCRITURA EN EXCEL (openpyxl + filelock) ============================

def insertar_multiples_registros(ruta_archivo, hoja_objetivo, columnas, lista_datos, contrasena):
    """
    Inserta múltiples registros en el archivo Excel usando openpyxl.

    El acceso concurrente está protegido por un FileLock: si dos usuarios
    intentan guardar al mismo tiempo, el segundo espera a que el primero
    termine y libere el lock antes de proceder.

    Parámetros
    ----------
    ruta_archivo  : str   - Ruta completa al archivo .xlsm
    hoja_objetivo : str   - Nombre de la hoja donde se insertarán los datos
    columnas      : list  - Lista de índices de columna (base 1) donde escribir
    lista_datos   : list  - Lista de tuplas, una por registro a insertar
    contrasena    : str   - Contraseña para desproteger/proteger la hoja
    """

    # FileLock garantiza acceso exclusivo al archivo entre procesos e hilos.
    # timeout=-1 significa espera indefinida (no lanza error por timeout).
    lock = FileLock(RUTA_LOCK, timeout=-1)

    with lock:
        # Abrir el libro manteniendo macros VBA (keep_vba=True)
        wb = load_workbook(ruta_archivo, keep_vba=True)
        ws = wb[hoja_objetivo]

        # ── Desproteger hoja ────────────────────────────────────────────────
        ws.protection.sheet = False

        # ── Encontrar la última fila con datos en alguna de las columnas ────
        ultima_fila = 1
        for col in columnas:
            for row_idx in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=row_idx, column=col).value
                if cell_value not in (None, ""):
                    if row_idx > ultima_fila:
                        ultima_fila = row_idx

        # ── Insertar todos los registros a partir de la siguiente fila libre ─
        for offset, datos in enumerate(lista_datos):
            nueva_fila = ultima_fila + offset + 1
            for i, col in enumerate(columnas):
                ws.cell(row=nueva_fila, column=col).value = datos[i]

        # ── Proteger hoja con la misma contraseña ───────────────────────────
        ws.protection.set_password(contrasena)

        # ── Guardar y cerrar ────────────────────────────────────────────────
        wb.save(ruta_archivo)
        wb.close()

    # Liberar memoria explícitamente
    del wb
    gc.collect()


def ejecutar_guardar_multiples(registros, usuario_actual):
    """
    Determina la hoja de destino según el usuario y lanza la escritura
    en el archivo Excel.
    """

    mapa_hojas = {
        "jorgeeh":   "Analítica de Contraloría",
        "albertoc":  "Control de Operaciones",
        "oscardy":   "Administrativa",
        "doragc":    "Riesgos y Cumplimiento",
        "zrestrepo": "Impuestos",
        "anamr":     "Contabilidad",
    }
    hoja = mapa_hojas.get(usuario_actual, "Admin")

    insertar_multiples_registros(
        ruta_archivo=RUTA_ARCHIVO,
        hoja_objetivo=hoja,
        columnas=[1, 2, 4, 7, 8, 11],
        lista_datos=registros,
        contrasena=CONTRASENA_EXCEL,
    )


# ============================================ UI =================================================

def aplicar_css():
    with open(RUTA_CSS, mode="r", encoding="utf-8") as f:
        css = f.read()
    st.markdown(f"<style>{css}</style>", unsafe_allow_html=True)


def mostrar_login():
    st.set_page_config(
        page_title="GCO | Inicio de Sesión",
        page_icon=RUTA_ICON,
        layout="centered",
        initial_sidebar_state="expanded",
    )

    container = st.container()
    with container:
        st.markdown(
            """
            <div class="h2" style='text-align: center;'>
                <h2>Formulario Informe de Gerencia</h2>
            </div>
            """,
            unsafe_allow_html=True,
        )
        st.divider()

        with st.form("login_form"):
            col01, col02 = st.columns([5, 5])

            with col01:
                col1, col2, col3 = st.columns([12, 1, 1])
                with col1:
                    st.image(RUTA_BIENVENIDA_GCO, width="stretch")

            with col02:
                usuario    = st.text_input("Usuario:", placeholder="Ej: usuario",  icon=":material/account_circle:")
                contraseña = st.text_input("Pin:",     placeholder="Ej: 1234",     type="password", icon=":material/passkey:")
                enviar     = st.form_submit_button("Iniciar Sesión", icon=":material/login:")

            alertas = st.container()
            with alertas:
                if enviar:
                    if not usuario or not contraseña:
                        st.warning("Campos obligatorios")
                    else:
                        login_ok       = False
                        nombre_usuario = None

                        for nombre, datos_bd in CREDENCIALES.items():
                            if usuario == datos_bd[0] and contraseña == datos_bd[1]:
                                st.session_state.autenticado    = True
                                st.session_state.usuario_actual = usuario
                                st.session_state.nombre_usuario = nombre
                                login_ok       = True
                                nombre_usuario = nombre
                                break

                        if login_ok:
                            st.success(
                                f"Bienvenido(a) {nombre_usuario}. Inicio de sesión completo",
                                icon=":material/how_to_reg:"
                            )
                            time.sleep(1)
                            st.rerun()
                        else:
                            st.error("Usuario o contraseña incorrectos")


def mostrar_formulario():
    # Limpiar campos si se solicitó en la iteración anterior
    if st.session_state.get("limpiar_campos", False):
        st.session_state.año_input           = None
        st.session_state.mes_input           = None
        st.session_state.concepto_input      = None
        st.session_state.especificacion_input = None
        st.session_state.ciudad_input        = None
        st.session_state.valor_input         = 0
        st.session_state.limpiar_campos      = False

    lista_especificaciones, lista_ciudades, lista_concepto_nuevo, diccionario_unidades = parametros(
        AREAS.get(st.session_state.usuario_actual)
    )
    lista_especificaciones_ordenadas = sorted(lista_especificaciones, key=lambda x: x != "No aplica")
    lista_ciudades_ordenadas         = sorted(lista_ciudades,         key=lambda x: x != "No aplica")

    st.set_page_config(
        page_title="GCO | Contraloría",
        page_icon=RUTA_ICON,
        layout="wide",
        initial_sidebar_state="expanded",
    )

    # ── Cabecera: bienvenida + botón cerrar sesión ────────────────────────
    container = st.container()
    with container:
        div1, div2 = st.columns([8, 2])
        with div1:
            st.success(
                "_" * 10 + f"Bienvenido(a), {st.session_state.nombre_usuario} ✌️" + "_" * 10,
                icon=":material/how_to_reg:"
            )
        with div2:
            cerrar_sesion = st.button("Cerrar Sesión", width="stretch")

        if cerrar_sesion:
            @st.dialog("¿Está seguro(a) de cerrar sesión?")
            def ventana_cerrar_sesion():
                if st.button("Confirmar Cerrar Sesión", icon=":material/logout:"):
                    for key in list(st.session_state.keys()):
                        del st.session_state[key]
                    st.rerun()
            ventana_cerrar_sesion()

    st.markdown(
        f"""
        <div class="h2-form" style='text-align: center; border-radius: 30px;'>
            <h2>Formulario {AREAS.get(st.session_state.usuario_actual)}</h2>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.divider()

    # ── Multiselect de conceptos ──────────────────────────────────────────
    if "saltarse_sync" not in st.session_state:
        st.session_state.saltarse_sync = False
    if "conceptos_buffer" not in st.session_state:
        st.session_state.conceptos_buffer = []

    concepto = st.container()
    with concepto:
        conceptos_seleccionados = st.multiselect(
            label="Conceptos",
            options=lista_concepto_nuevo,
            placeholder="💡Por favor seleccione los conceptos...",
            key="select_concepto",
        )
        st.session_state.conceptos_buffer = conceptos_seleccionados
        sincronizar_filas(conceptos_seleccionados)

    # ── Cabecera de columnas ──────────────────────────────────────────────
    inputs_form = st.container()
    (
        concepto_select, año_select, mes_select,
        especificacion_select, ciudad_select, valor_select,
        guia, duplicar, eliminar
    ) = st.columns([4, 2.5, 2.665, 3, 3, 3, 4, 1, 1])

    año_actual  = datetime.now().year
    años        = [año_actual - 2, año_actual - 1, año_actual]
    mes_defecto = datetime.now().month - 1

    años_seleccionados          = []
    meses_seleccionados         = []
    especificaciones_seleccionadas = []
    ciudades_seleccionadas      = []
    valores_seleccionados       = []

    with inputs_form:
        if conceptos_seleccionados:
            with concepto_select:
                st.write("Conceptos")
            with año_select:
                st.write("Año")
            with mes_select:
                st.write("Mes")
            with especificacion_select:
                st.write("Especificación")
            with ciudad_select:
                st.write("Ciudad")
            with valor_select:
                st.write("Valores")
            with guia:
                st.write("Guía")

        if "filas" not in st.session_state:
            st.session_state.filas = []

        # ── Filas dinámicas ───────────────────────────────────────────────
        for idx, fila in enumerate(st.session_state.filas):
            concepto  = fila["concepto"]
            fila_id   = fila["id"]

            with concepto_select:
                st.badge(concepto)

            with año_select:
                año = st.selectbox(
                    label="Año",
                    label_visibility="collapsed",
                    options=años,
                    index=2,
                    key=f"select_año_{fila_id}",
                )
                años_seleccionados.append(año)

            with mes_select:
                mes = st.selectbox(
                    label="Mes",
                    label_visibility="collapsed",
                    options=MESES,
                    index=mes_defecto,
                    key=f"select_mes_{fila_id}",
                )
                meses_seleccionados.append(mes)

            with especificacion_select:
                especificacion = st.selectbox(
                    label="Especificacion",
                    label_visibility="collapsed",
                    options=lista_especificaciones_ordenadas,
                    index=0,
                    key=f"select_especificacion_{fila_id}",
                )
                especificaciones_seleccionadas.append(especificacion)

            with ciudad_select:
                ciudad = st.selectbox(
                    label="Ciudad",
                    label_visibility="collapsed",
                    options=lista_ciudades_ordenadas,
                    index=0,
                    key=f"select_ciudad_{fila_id}",
                )
                ciudades_seleccionadas.append(ciudad)

            with valor_select:
                tipo_valor = diccionario_unidades.get(concepto, "")

                if tipo_valor == "Cantidad":
                    valor = st.number_input(
                        label="Unidades:",
                        label_visibility="collapsed",
                        format="%d",
                        step=1,
                        key=f"select_valor_{fila_id}",
                        icon=":material/dataset:",
                    )
                    valores_seleccionados.append(valor)
                    valor_formateado = locale.format_string("%.0f", valor, grouping=True)
                    with guia:
                        st.badge(f"Unidades: {valor_formateado}")

                elif tipo_valor == "Pesos":
                    valor = st.number_input(
                        label="Valor",
                        label_visibility="collapsed",
                        format="%d",
                        step=1,
                        key=f"select_valor_{fila_id}",
                        icon=":material/payments:",
                    )
                    valores_seleccionados.append(valor)
                    valor_formateado = locale.format_string("%.0f", valor, grouping=True)
                    with guia:
                        st.badge(f"Valor: ${valor_formateado}")

                elif tipo_valor == "Porcentaje":
                    valor = st.number_input(
                        label="Porcentaje",
                        label_visibility="collapsed",
                        format="%f",
                        step=1.0,
                        key=f"select_valor_{fila_id}",
                        icon=":material/percent:",
                    )
                    valores_seleccionados.append(valor)
                    with guia:
                        st.badge(f"Porcentaje: {valor}%")

                elif tipo_valor == "Toneladas":
                    valor = st.number_input(
                        label="Toneladas:",
                        label_visibility="collapsed",
                        format="%f",
                        step=1.0,
                        key=f"select_valor_{fila_id}",
                        icon=":material/weight:",
                    )
                    valores_seleccionados.append(valor)
                    valor_formateado = locale.format_string("%.0f", valor, grouping=True)
                    with guia:
                        st.badge(f"Toneladas: {valor_formateado}")

            # ── Botón duplicar fila ───────────────────────────────────────
            with duplicar:
                if st.button(
                    label="",
                    key=f"dup_{fila_id}",
                    icon=":material/control_point_duplicate:",
                    width="stretch",
                    help="Duplicar",
                ):
                    st.session_state.filas.insert(
                        idx + 1,
                        {"concepto": concepto, "id": str(uuid.uuid4())}
                    )
                    st.rerun()

            # ── Botón eliminar fila ───────────────────────────────────────
            with eliminar:
                total_concepto = sum(
                    1 for f in st.session_state.filas
                    if f["concepto"] == concepto
                )
                puede_eliminar = total_concepto > 1

                if st.button(
                    label="",
                    key=f"del_{fila_id}",
                    icon=":material/delete:",
                    disabled=not puede_eliminar,
                    help="Eliminar" if puede_eliminar else "Para eliminar este concepto, deselecciónelo desde el multiselect",
                ) and puede_eliminar:
                    for k in list(st.session_state.keys()):
                        if k.endswith(fila_id):
                            del st.session_state[k]

                    st.session_state.filas = [
                        f for f in st.session_state.filas
                        if f["id"] != fila_id
                    ]
                    st.rerun()

    # ── Botón enviar todo ─────────────────────────────────────────────────
    total_registros = len(st.session_state.filas)
    if conceptos_seleccionados:
        btn_enviar = st.button(
            label=f"Enviar Todo ({total_registros})",
            key="btn_enviar",
            width="stretch",
            icon=":material/send:",
        )
        if btn_enviar:
            @st.dialog("¿Está seguro(a) de enviar los registros?")
            def ventana_enviar_todo():
                if st.button("Enviar registros"):
                    with st.spinner("Enviando registros..."):

                        # Construir lista de registros a guardar
                        registros = [
                            (
                                años_seleccionados[idx],
                                meses_seleccionados[idx],
                                fila["concepto"],
                                especificaciones_seleccionadas[idx],
                                ciudades_seleccionadas[idx],
                                valores_seleccionados[idx],
                            )
                            for idx, fila in enumerate(st.session_state.filas)
                        ]

                        # Lanzar escritura en hilo separado
                        hilo_guardar = threading.Thread(
                            target=ejecutar_guardar_multiples,
                            args=(registros, st.session_state.usuario_actual),
                        )
                        hilo_guardar.start()

                        # Barra de progreso mientras el hilo trabaja
                        barra_carga = st.progress(0)
                        progreso = 0
                        while hilo_guardar.is_alive():
                            progreso = min(progreso + 5, 85)
                            barra_carga.progress(progreso)
                            time.sleep(0.8)

                        barra_carga.progress(100)
                        time.sleep(0.2)
                        barra_carga.empty()

                        st.toast("Registros enviados y guardados con éxito", icon=":material/folder_check:")
                        st.success(f"Se enviaron {total_registros} registro(s)")

                        # Limpiar estado después de guardar
                        st.session_state.filas           = []
                        st.session_state.conceptos_buffer = []

                        for k in list(st.session_state.keys()):
                            if any(k.startswith(p) for p in (
                                "select_año_", "select_mes_",
                                "select_especificacion_", "select_ciudad_",
                                "select_valor_",
                            )):
                                del st.session_state[k]

                        st.session_state.select_concepto = []
                        time.sleep(0.4)
                        st.rerun()

            ventana_enviar_todo()


# ============================================ MAIN ===============================================

def main():
    aplicar_css()

    # Inicializar session_state
    defaults = {
        "autenticado":     False,
        "usuario_actual":  "",
        "registros_tabla": [],
        "limpiar_campos":  False,
        "saltarse_sync":   False,
        "select_concepto": [],
        "filas":           [],
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value

    st.sidebar.image(RUTA_IMAGE)

    if st.session_state.autenticado:
        mostrar_formulario()
    else:
        mostrar_login()

    st.divider()
    st.image(RUTA_ICON_MARCAS, width="stretch")


if __name__ == "__main__":
    main()